from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path("/home/ubuntu/upload/Спринт1_доработанный.xlsx")
OUTPUT_DIR = Path("/home/ubuntu/rent-bez-figma/docs/sprint-1-source")


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", "").split())


def escape_cell(value: str) -> str:
    return value.replace("|", "\\|")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(SOURCE, data_only=False)
    extracted: dict[str, list[list[str]]] = {}

    for worksheet in workbook.worksheets:
        rows = [
            [normalize(cell.value) for cell in row]
            for row in worksheet.iter_rows(values_only=False)
        ]
        while rows and not any(rows[-1]):
            rows.pop()
        extracted[worksheet.title] = rows

    (OUTPUT_DIR / "sprint-1-full.json").write_text(
        json.dumps(extracted, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    markdown_parts = ["# Полное извлечение источника Sprint 1\n"]
    for sheet_name, rows in extracted.items():
        markdown_parts.append(f"## {sheet_name}\n")
        if not rows:
            markdown_parts.append("Лист не содержит данных.\n")
            continue
        max_columns = max(len(row) for row in rows)
        normalized_rows = [row + [""] * (max_columns - len(row)) for row in rows]
        header = normalized_rows[0]
        markdown_parts.append("| " + " | ".join(escape_cell(item) for item in header) + " |")
        markdown_parts.append("| " + " | ".join("---" for _ in header) + " |")
        for row in normalized_rows[1:]:
            markdown_parts.append("| " + " | ".join(escape_cell(item) for item in row) + " |")
        markdown_parts.append("")

    (OUTPUT_DIR / "sprint-1-full.md").write_text(
        "\n".join(markdown_parts), encoding="utf-8"
    )
    print(f"Extracted {len(extracted)} worksheets to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
